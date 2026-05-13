"""
routes/export.py - Data export to CSV and Excel
"""
import csv
import io
from flask import Blueprint, request, send_file, flash, redirect, url_for
from routes.auth import login_required
import db as database

export_bp = Blueprint('export', __name__, url_prefix='/export')


def _build_query(table, q, filter_col, filter_val):
    allowed_tables = {'employees', 'products'}
    if table not in allowed_tables:
        raise ValueError('Invalid table.')

    conditions, params = [], []
    if q:
        if table == 'employees':
            conditions.append('(first_name LIKE %s OR last_name LIKE %s OR email LIKE %s)')
            like = f'%{q}%'
            params += [like, like, like]
        else:
            conditions.append('(name LIKE %s OR sku LIKE %s)')
            like = f'%{q}%'
            params += [like, like]
    if filter_val:
        conditions.append(f'{filter_col} = %s')
        params.append(filter_val)

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''
    return f'SELECT * FROM {table} {where} ORDER BY id', params


@export_bp.route('/csv/<table>')
@login_required
def export_csv(table):
    q          = request.args.get('q', '')
    filter_col = 'department' if table == 'employees' else 'category'
    filter_val = request.args.get(filter_col, '')

    try:
        sql, params = _build_query(table, q, filter_col, filter_val)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('dashboard.index'))

    try:
        with database.get_cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()
    except Exception as e:
        flash(f'Export error: {e}', 'danger')
        return redirect(url_for('dashboard.index'))

    if not rows:
        flash('No data to export.', 'warning')
        return redirect(url_for(f'{table}.index'))

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    for row in rows:
        writer.writerow({k: (str(v) if v is not None else '') for k, v in row.items()})

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'{table}_export.csv'
    )


@export_bp.route('/excel/<table>')
@login_required
def export_excel(table):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl is required for Excel export. Run: pip install openpyxl', 'danger')
        return redirect(url_for(f'{table}.index'))

    q          = request.args.get('q', '')
    filter_col = 'department' if table == 'employees' else 'category'
    filter_val = request.args.get(filter_col, '')

    try:
        sql, params = _build_query(table, q, filter_col, filter_val)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('dashboard.index'))

    try:
        with database.get_cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()
    except Exception as e:
        flash(f'Export error: {e}', 'danger')
        return redirect(url_for('dashboard.index'))

    if not rows:
        flash('No data to export.', 'warning')
        return redirect(url_for(f'{table}.index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table.capitalize()

    # Header row
    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace('_', ' ').title())
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='1a1a2e')
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(row[key]) if row[key] is not None else '')

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{table}_export.xlsx'
    )
